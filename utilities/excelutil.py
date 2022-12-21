import openpyxl

def getRowcount(file,sheetname):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetname]
    return sheet.max_row

def getColumncount(file,sheetname):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetname]
    return sheet.max_column

def readData(file,sheetname,rowno,colno):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.cell(rowno,colno).value

def writeData(file,sheetname,rowno,colno,data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    sheet.cell(rowno,colno).value=data
    wb.save(file)